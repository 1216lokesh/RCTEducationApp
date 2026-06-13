import os
import sys
import subprocess

# Ensure openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def create_excel_sheet(results=None, return_cases=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Enable grid lines visibility
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "Test Case ID", "Feature / Module", "Sub-feature", 
        "Test Case Description", "Pre-conditions", "Test Steps", 
        "Expected Result", "Actual Result", "Status (Pass/Fail)", "Priority"
    ]
    
    test_cases = [
        {
            "id": "TC-AUTH-01", "module": "Authentication", "sub": "Registration",
            "desc": "User registration with valid patient details",
            "pre": "Access to registration page",
            "steps": "1. Fill in First Name, Last Name, Email, Password, Phone, Language.\n2. Click Register button.",
            "expected": "Account created successfully. User redirected to login page. Password stored securely hashed in DB.",
            "priority": "High"
        },
        {
            "id": "TC-AUTH-02", "module": "Authentication", "sub": "Registration",
            "desc": "User registration with already registered email",
            "pre": "Email already registered in DB",
            "steps": "1. Attempt registration using an existing email.\n2. Submit registration form.",
            "expected": "Error message shows 'Email already registered'. Registration rejected.",
            "priority": "High"
        },
        {
            "id": "TC-AUTH-03", "module": "Authentication", "sub": "Registration",
            "desc": "User registration with invalid email format",
            "pre": "Access to registration page",
            "steps": "1. Fill in registration form with invalid email format (e.g. 'invalid-email').\n2. Submit.",
            "expected": "Validation error displayed requesting a valid email address.",
            "priority": "Medium"
        },
        {
            "id": "TC-AUTH-04", "module": "Authentication", "sub": "Login",
            "desc": "Login with valid credentials (Patient)",
            "pre": "Valid patient account exists",
            "steps": "1. Enter valid email and password.\n2. Click Login.",
            "expected": "Login successful. User redirected to Patient Dashboard. Session initiated.",
            "priority": "High"
        },
        {
            "id": "TC-AUTH-05", "module": "Authentication", "sub": "Login",
            "desc": "Login with invalid credentials",
            "pre": "Any user state",
            "steps": "1. Enter invalid email or password.\n2. Click Login.",
            "expected": "Login fails. Error message shows 'Login failed' or 'Invalid credentials'.",
            "priority": "High"
        },
        {
            "id": "TC-AUTH-06", "module": "Authentication", "sub": "Login",
            "desc": "Remember Me functionality",
            "pre": "Access to login page",
            "steps": "1. Enter valid credentials.\n2. Check 'Remember Me' box.\n3. Login.\n4. Close and reopen browser.\n5. Navigate to app URL.",
            "expected": "User session is restored automatically without login page redirect using remember_token cookie.",
            "priority": "Medium"
        },
        {
            "id": "TC-AUTH-07", "module": "Authentication", "sub": "Logout",
            "desc": "User logout",
            "pre": "User is logged in",
            "steps": "1. Click Logout button/link.",
            "expected": "Session destroyed. Cookies cleared. User redirected to Login page.",
            "priority": "High"
        },
        {
            "id": "TC-AUTH-08", "module": "Authentication", "sub": "Forgot Password",
            "desc": "Forgot password OTP request",
            "pre": "Access to forgot password page",
            "steps": "1. Enter registered email address.\n2. Click 'Send OTP'.",
            "expected": "OTP is sent to user email. Redirected to OTP verification screen.",
            "priority": "High"
        },
        {
            "id": "TC-AUTH-09", "module": "Authentication", "sub": "Forgot Password",
            "desc": "Verify OTP and reset password",
            "pre": "OTP has been sent to user",
            "steps": "1. Enter valid OTP.\n2. Enter new password and confirm it.\n3. Click Reset.",
            "expected": "Password updated in database. User can now login with new password.",
            "priority": "High"
        },
        {
            "id": "TC-PAT-01", "module": "Patient Journey", "sub": "Dashboard",
            "desc": "Access Patient Dashboard",
            "pre": "User is logged in as Patient",
            "steps": "1. Complete login as patient.\n2. View dashboard.",
            "expected": "Dashboard loaded. Shows patient's name, active procedure info, and educational status.",
            "priority": "High"
        },
        {
            "id": "TC-PAT-02", "module": "Patient Journey", "sub": "Consent",
            "desc": "Fill and submit patient consent form",
            "pre": "Patient is logged in; consent not yet signed",
            "steps": "1. Navigate to Consent form.\n2. Check consent boxes and sign.\n3. Submit.",
            "expected": "Consent status updated to signed in database. Next step (Baseline Questionnaire) unlocked.",
            "priority": "High"
        },
        {
            "id": "TC-PAT-03", "module": "Patient Journey", "sub": "Baseline Questionnaire",
            "desc": "Complete Baseline Questionnaire",
            "pre": "Patient is logged in; consent signed",
            "steps": "1. Navigate to Baseline Questionnaire.\n2. Answer all questions.\n3. Submit.",
            "expected": "Anxiety Survey unlocked. Answers recorded in database.",
            "priority": "High"
        },
        {
            "id": "TC-PAT-04", "module": "Patient Journey", "sub": "Anxiety Survey",
            "desc": "Complete Anxiety Survey",
            "pre": "Patient is logged in; baseline questionnaire completed",
            "steps": "1. Navigate to Anxiety Survey.\n2. Complete survey.\n3. Submit.",
            "expected": "Educational Content unlocked. Scores saved in database.",
            "priority": "High"
        },
        {
            "id": "TC-PAT-05", "module": "Patient Journey", "sub": "Educational Content",
            "desc": "View educational videos/documents",
            "pre": "Patient is logged in; anxiety survey completed",
            "steps": "1. Navigate to Education/Counselling page.\n2. View procedure video and reading materials.",
            "expected": "Viewing progress recorded. Quiz section unlocked.",
            "priority": "High"
        },
        {
            "id": "TC-PAT-06", "module": "Patient Journey", "sub": "Quiz",
            "desc": "Take knowledge quiz on procedure",
            "pre": "Patient is logged in; educational content viewed",
            "steps": "1. Navigate to Quiz page.\n2. Answer multiple choice questions.\n3. Submit.",
            "expected": "Quiz score calculated and stored. Postop Questionnaire unlocked. Pass/fail status logged.",
            "priority": "High"
        },
        {
            "id": "TC-PAT-07", "module": "Patient Journey", "sub": "Post-op Questionnaire",
            "desc": "Submit post-operative feedback",
            "pre": "Patient is logged in; quiz completed",
            "steps": "1. Navigate to Postop Questionnaire.\n2. Enter feedback.\n3. Submit.",
            "expected": "Feedback saved. Follow-up instructions displayed.",
            "priority": "Medium"
        },
        {
            "id": "TC-PAT-08", "module": "Patient Journey", "sub": "Satisfaction Survey",
            "desc": "Submit patient satisfaction survey",
            "pre": "Patient is logged in; postop completed",
            "steps": "1. Navigate to Satisfaction Survey.\n2. Grade the portal experience.\n3. Submit.",
            "expected": "Satisfaction rating saved. Portal journey completed.",
            "priority": "Medium"
        },
        {
            "id": "TC-ADM-01", "module": "Admin Controls", "sub": "Dashboard",
            "desc": "View Admin Dashboard",
            "pre": "User logged in as Admin",
            "steps": "1. Complete login as admin.\n2. View dashboard.",
            "expected": "Admin dashboard loaded. Displays active patient count, quiz completion rates, and system overview.",
            "priority": "High"
        },
        {
            "id": "TC-ADM-02", "module": "Admin Controls", "sub": "Patients List",
            "desc": "Search and filter patients",
            "pre": "Admin logged in",
            "steps": "1. Navigate to Patients list.\n2. Search patient by name or email.\n3. Filter by procedure or status.",
            "expected": "List updates dynamically to match search query and filters.",
            "priority": "High"
        },
        {
            "id": "TC-ADM-03", "module": "Admin Controls", "sub": "Patient Detail",
            "desc": "View detailed patient scores and answers",
            "pre": "Admin logged in",
            "steps": "1. Click on a patient in the list.",
            "expected": "Patient detail page loaded showing baseline, anxiety, quiz, postop, and satisfaction scores.",
            "priority": "High"
        },
        {
            "id": "TC-ADM-04", "module": "Admin Controls", "sub": "Attendance",
            "desc": "Record patient attendance",
            "pre": "Admin logged in",
            "steps": "1. Navigate to Attendance page.\n2. Check attendance checkbox for patient.\n3. Save.",
            "expected": "Attendance status saved in DB and updated on screen.",
            "priority": "Medium"
        },
        {
            "id": "TC-ADM-05", "module": "Admin Controls", "sub": "Procedure Assignment",
            "desc": "Assign procedure to patient",
            "pre": "Admin logged in",
            "steps": "1. Open patient details.\n2. Select procedure from dropdown.\n3. Click Assign.",
            "expected": "Procedure assigned to patient. Patient's portal dashboard updates accordingly.",
            "priority": "High"
        },
        {
            "id": "TC-ADM-06", "module": "Admin Controls", "sub": "Data Export",
            "desc": "Export patient data to Excel/CSV",
            "pre": "Admin logged in",
            "steps": "1. Navigate to Export page.\n2. Click 'Export Patients Data' or 'Export Scores'.",
            "expected": "Excel/CSV file downloaded containing corresponding patient and scoring data.",
            "priority": "High"
        },
        {
            "id": "TC-SEC-01", "module": "Security & RBAC", "sub": "Access Control",
            "desc": "Access patient dashboard without login",
            "pre": "User is unauthenticated",
            "steps": "1. Try to directly navigate to '/patient/dashboard.php'.",
            "expected": "User redirected to login page (302 Redirect).",
            "priority": "High"
        },
        {
            "id": "TC-SEC-02", "module": "Security & RBAC", "sub": "Access Control",
            "desc": "Access admin dashboard as Patient",
            "pre": "User logged in as Patient",
            "steps": "1. Navigate to '/admin/dashboard.php'.",
            "expected": "Access Denied (403 Forbidden) page displayed.",
            "priority": "High"
        },
        {
            "id": "TC-SEC-03", "module": "Security & RBAC", "sub": "Access Control",
            "desc": "Access admin API endpoint as Patient",
            "pre": "User logged in as Patient",
            "steps": "1. Send request to '/api/admin/get_patients.php'.",
            "expected": "API returns 403 Forbidden.",
            "priority": "High"
        },
        {
            "id": "TC-SEC-04", "module": "Security & RBAC", "sub": "SQL Injection",
            "desc": "SQL Injection prevention on Login",
            "pre": "Any user state",
            "steps": "1. Try login with email: \"admin' OR '1'='1\" and password: \"foo\".",
            "expected": "Login fails. SQL execution handles parameters safely (via escaping or prepared statements).",
            "priority": "High"
        }
    ]
    
    if return_cases:
        return test_cases
    
    # Styles definition
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    # Zebra striping
    even_row_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Fonts
    body_font = Font(name="Segoe UI", size=10)
    bold_body_font = Font(name="Segoe UI", size=10, bold=True)
    
    # Priority coloring
    high_priority_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # Light red
    med_priority_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid") # Light yellow
    
    # Status coloring
    pass_fill = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid") # Light green
    fail_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # Light red
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Write Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_all
    
    ws.row_dimensions[1].height = 28
    
    # Write Data
    for row_num, tc in enumerate(test_cases, 2):
        row_fill = even_row_fill if row_num % 2 == 0 else white_fill
        
        # ID
        c_id = ws.cell(row=row_num, column=1, value=tc["id"])
        c_id.font = bold_body_font
        c_id.alignment = align_center
        c_id.fill = row_fill
        c_id.border = border_all
        
        # Module
        c_mod = ws.cell(row=row_num, column=2, value=tc["module"])
        c_mod.font = body_font
        c_mod.alignment = align_left
        c_mod.fill = row_fill
        c_mod.border = border_all
        
        # Sub
        c_sub = ws.cell(row=row_num, column=3, value=tc["sub"])
        c_sub.font = body_font
        c_sub.alignment = align_left
        c_sub.fill = row_fill
        c_sub.border = border_all
        
        # Desc
        c_desc = ws.cell(row=row_num, column=4, value=tc["desc"])
        c_desc.font = body_font
        c_desc.alignment = align_left
        c_desc.fill = row_fill
        c_desc.border = border_all
        
        # Pre
        c_pre = ws.cell(row=row_num, column=5, value=tc["pre"])
        c_pre.font = body_font
        c_pre.alignment = align_left
        c_pre.fill = row_fill
        c_pre.border = border_all
        
        # Steps
        c_steps = ws.cell(row=row_num, column=6, value=tc["steps"])
        c_steps.font = body_font
        c_steps.alignment = align_left
        c_steps.fill = row_fill
        c_steps.border = border_all
        
        # Expected
        c_exp = ws.cell(row=row_num, column=7, value=tc["expected"])
        c_exp.font = body_font
        c_exp.alignment = align_left
        c_exp.fill = row_fill
        c_exp.border = border_all
        
        # Actual Result
        actual_val = results.get(tc["id"], {}).get("actual", "") if results else ""
        c_act = ws.cell(row=row_num, column=8, value=actual_val)
        c_act.font = body_font
        c_act.alignment = align_left
        c_act.fill = row_fill
        c_act.border = border_all
        
        # Status
        status_val = results.get(tc["id"], {}).get("status", "") if results else ""
        c_stat = ws.cell(row=row_num, column=9, value=status_val)
        c_stat.font = bold_body_font if status_val else body_font
        c_stat.alignment = align_center
        c_stat.border = border_all
        if status_val == "Pass":
            c_stat.fill = pass_fill
        elif status_val == "Fail":
            c_stat.fill = fail_fill
        else:
            c_stat.fill = row_fill
        
        # Priority
        p_val = tc["priority"]
        c_prio = ws.cell(row=row_num, column=10, value=p_val)
        c_prio.font = bold_body_font
        c_prio.alignment = align_center
        c_prio.border = border_all
        if p_val == "High":
            c_prio.fill = high_priority_fill
        elif p_val == "Medium":
            c_prio.fill = med_priority_fill
        else:
            c_prio.fill = row_fill
            
        # Add custom row height for better spacing
        ws.row_dimensions[row_num].height = 45

    # Define Column Widths
    col_widths = {
        "A": 15,  # Test Case ID
        "B": 22,  # Feature / Module
        "C": 25,  # Sub-feature
        "D": 38,  # Test Case Description
        "E": 32,  # Pre-conditions
        "F": 48,  # Test Steps
        "G": 48,  # Expected Result
        "H": 30,  # Actual Result
        "I": 18,  # Status
        "J": 12   # Priority
    }
    
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save
    excel_path = os.path.join(os.path.dirname(__file__), "test_cases.xlsx")
    try:
        wb.save(excel_path)
        print(f"Excel file created successfully at: {excel_path}")
    except PermissionError:
        alt_path = os.path.join(os.path.dirname(__file__), "test_cases_report.xlsx")
        wb.save(alt_path)
        print(f"\nWARNING: Could not write to '{excel_path}' because it is open in Excel.")
        print(f"Fallback saved successfully at: {alt_path}")

if __name__ == "__main__":
    create_excel_sheet()
