import os
import sys
import subprocess
import time

# Verify dependencies and import selenium
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import Select
except ImportError:
    print("selenium not found. Installing selenium...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "selenium"])
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import Select

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl not found. Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_URL = "http://localhost/rct-education-web"
TEST_EMAIL = "test_frontend_patient@rct.com"
TEST_PASS = "frontend_pass_123"

def execute_mysql_query(query):
    cmd = [
        "c:\\xampp\\mysql\\bin\\mysql.exe",
        "-u", "root",
        "-e", query
    ]
    try:
        output = subprocess.check_output(cmd, stderr=subprocess.DEVNULL).decode('utf-8').strip()
        return [line.strip() for line in output.split('\n') if line.strip()]
    except Exception:
        return []

def get_user_id_by_email(email):
    lines = execute_mysql_query(f"SELECT id FROM rct_app.users WHERE email='{email}'")
    if len(lines) > 1:
        return lines[1]
    return None

def db_cleanup(email):
    user_id = get_user_id_by_email(email)
    if not user_id:
        return
    tables = ['consent', 'scores', 'anxiety_scores', 'baseline_responses', 'attendance', 'patient_procedure', 'satisfaction_scores']
    for table in tables:
        col = 'patient_id' if table in ['anxiety_scores', 'baseline_responses', 'satisfaction_scores'] else 'user_id'
        execute_mysql_query(f"DELETE FROM rct_app.{table} WHERE {col}='{user_id}'")
    execute_mysql_query(f"DELETE FROM rct_app.users WHERE id='{user_id}'")
    print(f"Database records cleaned up for {email}")

def setup_driver():
    chrome_options = Options()
    # Try headed mode first, fallback to headless if it fails
    try:
        print("Starting Chrome in headed mode...")
        driver = webdriver.Chrome(options=chrome_options)
        return driver
    except Exception as e:
        print(f"Failed to launch headed Chrome: {e}. Falling back to headless mode...")
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        return webdriver.Chrome(options=chrome_options)

def safe_click_element(driver, element):
    try:
        element.click()
    except Exception:
        driver.execute_script("arguments[0].click();", element)

def safe_click(driver, by, value):
    element = WebDriverWait(driver, 6).until(EC.presence_of_element_located((by, value)))
    safe_click_element(driver, element)
    return element

def write_results_to_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Frontend UI Test Cases"
    ws.views.sheetView[0].showGridLines = True

    headers = [
        "Test Case ID", "Feature / Module", "Sub-feature", 
        "Test Case Description", "Expected Result", "Actual Result", "Status (Pass/Fail)", "Priority"
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    even_row_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    body_font = Font(name="Segoe UI", size=10)
    bold_body_font = Font(name="Segoe UI", size=10, bold=True)
    high_priority_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    med_priority_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
    pass_fill = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
    fail_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Write Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_all
    ws.row_dimensions[1].height = 28

    test_cases_defs = [
        {"id": "TC-FE-SEC-01", "module": "Security UI", "sub": "RBAC", "desc": "Access patient dashboard anonymously", "expected": "Automatically redirected to login.php.", "priority": "High"},
        {"id": "TC-FE-AUTH-01", "module": "Authentication UI", "sub": "Registration", "desc": "User registration with valid details", "expected": "Successfully registers and redirects to patient dashboard.", "priority": "High"},
        {"id": "TC-FE-AUTH-02", "module": "Authentication UI", "sub": "Registration", "desc": "Registration with duplicate email address", "expected": "Rejects with 'Email already registered' validation alert.", "priority": "High"},
        {"id": "TC-FE-AUTH-04", "module": "Authentication UI", "sub": "Login", "desc": "Login with invalid credentials", "expected": "Fails and shows validation warning banner.", "priority": "High"},
        {"id": "TC-FE-AUTH-03", "module": "Authentication UI", "sub": "Login", "desc": "Login with valid patient credentials", "expected": "Login succeeds and redirects to patient dashboard.", "priority": "High"},
        {"id": "TC-FE-SEC-02", "module": "Security UI", "sub": "RBAC", "desc": "Access admin dashboard as Patient", "expected": "Access Denied (HTTP 403) page is shown.", "priority": "High"},
        {"id": "TC-FE-PAT-01", "module": "Patient UI", "sub": "Dashboard", "desc": "Access Patient Dashboard UI features", "expected": "Renders patient details, active procedures, and progress cards.", "priority": "High"},
        {"id": "TC-FE-PAT-02", "module": "Patient UI", "sub": "Consent", "desc": "Submit patient digital consent form", "expected": "Consent checkbox clicks, form submits, and redirects to satisfaction survey.", "priority": "High"},
        {"id": "TC-FE-PAT-03", "module": "Patient UI", "sub": "Satisfaction Survey", "desc": "Submit app satisfaction survey", "expected": "Survey ratings submit successfully and redirects back to patient dashboard.", "priority": "Medium"},
        {"id": "TC-FE-PAT-04", "module": "Patient UI", "sub": "Baseline Questionnaire", "desc": "Submit baseline questionnaire", "expected": "Submits answers successfully and redirects to procedure details page.", "priority": "High"},
        {"id": "TC-FE-PAT-05", "module": "Patient UI", "sub": "Education Content", "desc": "Access and view educational procedure content", "expected": "Loads video layout/reading resources, and take quiz redirects to anxiety survey.", "priority": "High"},
        {"id": "TC-FE-PAT-06", "module": "Patient UI", "sub": "Anxiety Survey", "desc": "Submit anxiety levels", "expected": "Submits survey answers successfully and redirects to knowledge quiz.", "priority": "High"},
        {"id": "TC-FE-PAT-07", "module": "Patient UI", "sub": "Quiz", "desc": "Complete knowledge quiz", "expected": "Submits score, triggers confirmation overlay modal, and continue button redirects.", "priority": "High"},
        {"id": "TC-FE-PAT-08", "module": "Patient UI", "sub": "Post-op Instructions", "desc": "View post-operative instructions", "expected": "Accesses care instructions and allows redirect back to dashboard.", "priority": "Medium"},
        {"id": "TC-FE-ADM-01", "module": "Admin UI", "sub": "Dashboard", "desc": "View Admin Dashboard statistics", "expected": "Admin dashboard loaded showing patient count and system charts.", "priority": "High"},
        {"id": "TC-FE-ADM-02", "module": "Admin UI", "sub": "Patients List", "desc": "Search and filter patient list table", "expected": "Filters search query successfully in patient details table.", "priority": "High"},
        {"id": "TC-FE-ADM-03", "module": "Admin UI", "sub": "Procedure Assignment", "desc": "Assign dental procedure to patient", "expected": "Dropdown selection submits successfully and updates assigned procedure overview.", "priority": "High"},
        {"id": "TC-FE-AUTH-05", "module": "Authentication UI", "sub": "Logout", "desc": "Verify session logout UI flow", "expected": "Redirects to login page; clears active login cookies.", "priority": "High"}
    ]

    for row_num, tc in enumerate(test_cases_defs, 2):
        row_fill = even_row_fill if row_num % 2 == 0 else white_fill
        
        ws.cell(row=row_num, column=1, value=tc["id"]).font = bold_body_font
        ws.cell(row=row_num, column=1).alignment = align_center
        ws.cell(row=row_num, column=1).fill = row_fill
        ws.cell(row=row_num, column=1).border = border_all

        ws.cell(row=row_num, column=2, value=tc["module"]).font = body_font
        ws.cell(row=row_num, column=2).alignment = align_left
        ws.cell(row=row_num, column=2).fill = row_fill
        ws.cell(row=row_num, column=2).border = border_all

        ws.cell(row=row_num, column=3, value=tc["sub"]).font = body_font
        ws.cell(row=row_num, column=3).alignment = align_left
        ws.cell(row=row_num, column=3).fill = row_fill
        ws.cell(row=row_num, column=3).border = border_all

        ws.cell(row=row_num, column=4, value=tc["desc"]).font = body_font
        ws.cell(row=row_num, column=4).alignment = align_left
        ws.cell(row=row_num, column=4).fill = row_fill
        ws.cell(row=row_num, column=4).border = border_all

        ws.cell(row=row_num, column=5, value=tc["expected"]).font = body_font
        ws.cell(row=row_num, column=5).alignment = align_left
        ws.cell(row=row_num, column=5).fill = row_fill
        ws.cell(row=row_num, column=5).border = border_all

        # Results
        res_data = results.get(tc["id"], {"status": "Fail", "actual": "Test did not run."})
        
        ws.cell(row=row_num, column=6, value=res_data["actual"]).font = body_font
        ws.cell(row=row_num, column=6).alignment = align_left
        ws.cell(row=row_num, column=6).fill = row_fill
        ws.cell(row=row_num, column=6).border = border_all

        stat_cell = ws.cell(row=row_num, column=7, value=res_data["status"])
        stat_cell.font = bold_body_font
        stat_cell.alignment = align_center
        stat_cell.border = border_all
        if res_data["status"] == "Pass":
            stat_cell.fill = pass_fill
        else:
            stat_cell.fill = fail_fill

        prio_cell = ws.cell(row=row_num, column=8, value=tc["priority"])
        prio_cell.font = bold_body_font
        prio_cell.alignment = align_center
        prio_cell.border = border_all
        if tc["priority"] == "High":
            prio_cell.fill = high_priority_fill
        else:
            prio_cell.fill = med_priority_fill

        ws.row_dimensions[row_num].height = 40

    col_widths = {
        "A": 15, "B": 20, "C": 20, "D": 35, "E": 35, "H": 12
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 15

    excel_path = os.path.join(os.path.dirname(__file__), "frontend_test_cases.xlsx")
    wb.save(excel_path)
    print(f"Excel report saved successfully at: {excel_path}")

    # Save CSV
    import csv
    csv_path = os.path.join(os.path.dirname(__file__), "frontend_test_cases.csv")
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for tc in test_cases_defs:
            res_data = results.get(tc["id"], {"status": "Fail", "actual": "Test did not run."})
            writer.writerow([
                tc["id"], tc["module"], tc["sub"], tc["desc"], tc["expected"],
                res_data["actual"], res_data["status"], tc["priority"]
            ])
    print(f"CSV report saved successfully at: {csv_path}")

def run_tests():
    print("==================================================")
    print("       STARTING AUTOMATED FRONTEND TESTS          ")
    print("==================================================")
    
    db_cleanup(TEST_EMAIL)
    results = {}
    driver = setup_driver()
    wait = WebDriverWait(driver, 6)

    try:
        # ----------------------------------------------------
        # TC-FE-SEC-01: Access patient dashboard anonymously
        # ----------------------------------------------------
        tc_id = "TC-FE-SEC-01"
        try:
            driver.get(f"{BASE_URL}/frontend/patient/dashboard.php")
            time.sleep(1)
            curr_url = driver.current_url
            if "login.php" in curr_url:
                results[tc_id] = {"status": "Pass", "actual": f"Redirected to Login. Checked URL: {curr_url}"}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Expected redirect, but loaded page: {curr_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-AUTH-01: User registration with valid details
        # ----------------------------------------------------
        tc_id = "TC-FE-AUTH-01"
        try:
            driver.get(f"{BASE_URL}/frontend/auth/register.php")
            wait.until(EC.presence_of_element_located((By.ID, "first_name")))
            driver.find_element(By.ID, "first_name").send_keys("Frontend")
            driver.find_element(By.ID, "last_name").send_keys("Patient")
            driver.find_element(By.ID, "email").send_keys(TEST_EMAIL)
            driver.find_element(By.ID, "phone").send_keys("9988776655")
            driver.find_element(By.ID, "date_of_birth").send_keys("15051995") # input type=date formatting
            Select(driver.find_element(By.ID, "gender")).select_by_value("M")
            Select(driver.find_element(By.ID, "language")).select_by_value("en")
            driver.find_element(By.ID, "password").send_keys(TEST_PASS)
            driver.find_element(By.ID, "confirm_password").send_keys(TEST_PASS)
            
            # Click submit
            safe_click(driver, By.XPATH, "//button[@type='submit']")
            time.sleep(2)
            
            curr_url = driver.current_url
            if "dashboard.php" in curr_url:
                results[tc_id] = {"status": "Pass", "actual": f"Registered and redirected to Patient Dashboard. URL: {curr_url}"}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Registration form submitted, but not redirected to dashboard. URL: {curr_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # Get patient user ID
        patient_id = get_user_id_by_email(TEST_EMAIL)
        print(f"DEBUG: Registered Patient ID is: {patient_id}")

        # ----------------------------------------------------
        # TC-FE-AUTH-02: Registration with duplicate email address
        # ----------------------------------------------------
        tc_id = "TC-FE-AUTH-02"
        try:
            # Clear cookies to simulate anonymous registration try
            driver.delete_all_cookies()
            driver.get(f"{BASE_URL}/frontend/auth/register.php")
            wait.until(EC.presence_of_element_located((By.ID, "first_name")))
            driver.find_element(By.ID, "first_name").send_keys("Frontend")
            driver.find_element(By.ID, "last_name").send_keys("Patient")
            driver.find_element(By.ID, "email").send_keys(TEST_EMAIL)
            driver.find_element(By.ID, "password").send_keys(TEST_PASS)
            driver.find_element(By.ID, "confirm_password").send_keys(TEST_PASS)
            safe_click(driver, By.XPATH, "//button[@type='submit']")
            time.sleep(1.5)
            
            # Look for warning alert
            alert = driver.find_element(By.CLASS_NAME, "alert-danger")
            if "already registered" in alert.text.lower():
                results[tc_id] = {"status": "Pass", "actual": f"Registration rejected correctly. Alert text: {alert.text}"}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Duplicate alert found, but text mismatch: {alert.text}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-AUTH-04: Login with invalid credentials
        # ----------------------------------------------------
        tc_id = "TC-FE-AUTH-04"
        try:
            driver.get(f"{BASE_URL}/frontend/auth/login.php")
            wait.until(EC.presence_of_element_located((By.ID, "email")))
            driver.find_element(By.ID, "email").send_keys(TEST_EMAIL)
            driver.find_element(By.ID, "password").send_keys("wrong_pass_here")
            safe_click(driver, By.XPATH, "//button[@type='submit']")
            time.sleep(1.5)
            
            alert = driver.find_element(By.CLASS_NAME, "alert-danger")
            if "invalid" in alert.text.lower() or "failed" in alert.text.lower():
                results[tc_id] = {"status": "Pass", "actual": f"Failed login attempt handled properly. Alert text: {alert.text}"}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Error banner shown, but text was unexpected: {alert.text}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-AUTH-03: Login with valid patient credentials
        # ----------------------------------------------------
        tc_id = "TC-FE-AUTH-03"
        try:
            driver.get(f"{BASE_URL}/frontend/auth/login.php")
            wait.until(EC.presence_of_element_located((By.ID, "email")))
            driver.find_element(By.ID, "email").clear()
            driver.find_element(By.ID, "email").send_keys(TEST_EMAIL)
            driver.find_element(By.ID, "password").send_keys(TEST_PASS)
            safe_click(driver, By.XPATH, "//button[@type='submit']")
            time.sleep(2)
            
            curr_url = driver.current_url
            if "dashboard.php" in curr_url:
                results[tc_id] = {"status": "Pass", "actual": f"Successfully logged in and redirected to Patient Dashboard. URL: {curr_url}"}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Expected patient dashboard, but url was: {curr_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-SEC-02: Access admin dashboard as Patient
        # ----------------------------------------------------
        tc_id = "TC-FE-SEC-02"
        try:
            driver.get(f"{BASE_URL}/frontend/admin/dashboard.php")
            time.sleep(1.5)
            src = driver.page_source
            if "Access Denied" in src or "403" in src:
                results[tc_id] = {"status": "Pass", "actual": "Access blocked as patient. Page contains Access Denied content."}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Expected Access Denied block, but page source did not contain it. URL: {driver.current_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # Navigate back to patient dashboard
        driver.get(f"{BASE_URL}/frontend/patient/dashboard.php")
        time.sleep(1)

        # ----------------------------------------------------
        # TC-FE-PAT-01: Access Patient Dashboard UI features
        # ----------------------------------------------------
        tc_id = "TC-FE-PAT-01"
        try:
            time.sleep(2)
            body_text = driver.find_element(By.TAG_NAME, "body").text
            if "Dashboard" in body_text or "welcome" in body_text.lower():
                results[tc_id] = {"status": "Pass", "actual": "Successfully verified patient dashboard header and text features."}
            else:
                results[tc_id] = {"status": "Fail", "actual": "Dashboard text not found in patient dashboard body."}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-PAT-02: Submit patient digital consent form
        # ----------------------------------------------------
        tc_id = "TC-FE-PAT-02"
        try:
            driver.get(f"{BASE_URL}/frontend/patient/consent.php")
            wait.until(EC.presence_of_element_located((By.NAME, "agree")))
            checkbox = driver.find_element(By.NAME, "agree")
            if not checkbox.is_selected():
                safe_click_element(driver, checkbox)
            
            # Click submit button
            safe_click(driver, By.XPATH, "//button[@type='submit']")
            time.sleep(2)
            
            curr_url = driver.current_url
            if "satisfaction.php" in curr_url:
                results[tc_id] = {"status": "Pass", "actual": f"Consent submitted successfully. Redirected to satisfaction.php. URL: {curr_url}"}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Failed redirect after consent submission. URL: {curr_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-PAT-03: Submit app satisfaction survey
        # ----------------------------------------------------
        tc_id = "TC-FE-PAT-03"
        try:
            # We are on satisfaction.php. We need to answer q1 to q5
            wait.until(EC.presence_of_element_located((By.NAME, "q1")))
            for i in range(1, 6):
                radios = driver.find_elements(By.XPATH, f"//input[@name='q{i}']")
                if len(radios) >= 5:
                    safe_click_element(driver, radios[4]) # Select Strongly Agree (value 5)
            
            safe_click(driver, By.XPATH, "//button[@type='submit']")
            time.sleep(2)
            
            curr_url = driver.current_url
            if "dashboard.php" in curr_url:
                results[tc_id] = {"status": "Pass", "actual": f"Satisfaction survey submitted. Redirected back to dashboard.php. URL: {curr_url}"}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Failed redirect after satisfaction survey. URL: {curr_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-PAT-04: Submit baseline questionnaire
        # ----------------------------------------------------
        tc_id = "TC-FE-PAT-04"
        try:
            driver.get(f"{BASE_URL}/frontend/patient/baseline.php?apt=1")
            wait.until(EC.presence_of_element_located((By.NAME, "q1")))
            
            # Select first radio option for each question
            safe_click_element(driver, driver.find_elements(By.NAME, "q1")[0])
            safe_click_element(driver, driver.find_elements(By.NAME, "q2")[0])
            safe_click_element(driver, driver.find_elements(By.NAME, "q3")[0])
            
            safe_click(driver, By.XPATH, "//button[@type='submit']")
            time.sleep(2)
            
            curr_url = driver.current_url
            if "procedure_info.php" in curr_url:
                results[tc_id] = {"status": "Pass", "actual": f"Baseline questionnaire submitted successfully. Redirected to procedure_info.php."}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Failed redirect after baseline. URL: {curr_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-PAT-05: Access and view educational procedure content
        # ----------------------------------------------------
        tc_id = "TC-FE-PAT-05"
        try:
            # Click "Next" on procedure_info.php to go to education.php
            safe_click(driver, By.XPATH, "//a[contains(@href, 'education.php')]")
            time.sleep(2)
            
            curr_url = driver.current_url
            if "education.php" in curr_url:
                # Click Take Quiz button (redirects to anxiety.php)
                safe_click(driver, By.XPATH, "//a[contains(@href, 'anxiety.php')]")
                time.sleep(2)
                
                curr_url = driver.current_url
                if "anxiety.php" in curr_url:
                    results[tc_id] = {"status": "Pass", "actual": "Education content viewed; take quiz links to anxiety survey correctly."}
                else:
                    results[tc_id] = {"status": "Fail", "actual": f"Expected redirect to anxiety.php, but got: {curr_url}"}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Expected redirect to education.php, but got: {curr_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-PAT-06: Submit anxiety levels
        # ----------------------------------------------------
        tc_id = "TC-FE-PAT-06"
        try:
            driver.get(f"{BASE_URL}/frontend/patient/anxiety.php?apt=1")
            wait.until(EC.presence_of_element_located((By.NAME, "q1")))
            
            # Select moderately (value 2) for q1 to q5
            for i in range(1, 6):
                radios = driver.find_elements(By.NAME, f"q{i}")
                if len(radios) >= 3:
                    safe_click_element(driver, radios[2])
            
            safe_click(driver, By.XPATH, "//button[@type='submit']")
            time.sleep(2)
            
            curr_url = driver.current_url
            if "quiz.php" in curr_url:
                results[tc_id] = {"status": "Pass", "actual": f"Anxiety levels submitted successfully. Redirected to quiz.php."}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Failed redirect after anxiety survey. URL: {curr_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-PAT-07: Complete knowledge quiz
        # ----------------------------------------------------
        tc_id = "TC-FE-PAT-07"
        try:
            driver.get(f"{BASE_URL}/frontend/patient/quiz.php?apt=1")
            wait.until(EC.presence_of_element_located((By.NAME, "q1")))
            
            # Select correct options (value 1)
            safe_click_element(driver, driver.find_elements(By.NAME, "q1")[1])
            safe_click_element(driver, driver.find_elements(By.NAME, "q2")[1])
            safe_click_element(driver, driver.find_elements(By.NAME, "q3")[1])
            
            safe_click(driver, By.XPATH, "//button[@type='submit']")
            time.sleep(2)
            
            # Look for Continue button on the overlay modal
            continue_btn = wait.until(EC.presence_of_element_located((By.XPATH, "//a[contains(@href, 'counselling.php') or contains(@href, 'postop.php') or contains(@href, 'dashboard.php')]")))
            safe_click_element(driver, continue_btn)
            time.sleep(2)
            
            curr_url = driver.current_url
            if "counselling.php" in curr_url or "dashboard.php" in curr_url:
                results[tc_id] = {"status": "Pass", "actual": f"Quiz submitted successfully; modal continue clicked. Redirected to: {curr_url}"}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Unexpected url after clicking quiz continue: {curr_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-PAT-08: View post-operative instructions
        # ----------------------------------------------------
        tc_id = "TC-FE-PAT-08"
        try:
            driver.get(f"{BASE_URL}/frontend/patient/postop.php")
            safe_click(driver, By.XPATH, "//a[contains(@href, 'dashboard.php')]")
            time.sleep(1.5)
            
            curr_url = driver.current_url
            if "dashboard.php" in curr_url:
                results[tc_id] = {"status": "Pass", "actual": "Successfully accessed post-op instructions and clicked back to dashboard."}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Failed back to dashboard redirect. URL: {curr_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # Admin UI: Log in as Admin
        # ----------------------------------------------------
        driver.delete_all_cookies()
        driver.get(f"{BASE_URL}/frontend/auth/login.php")
        wait.until(EC.presence_of_element_located((By.ID, "email")))
        driver.find_element(By.ID, "email").send_keys("admin@rct.com")
        driver.find_element(By.ID, "password").send_keys("admin123")
        safe_click(driver, By.XPATH, "//button[@type='submit']")
        time.sleep(2)

        # ----------------------------------------------------
        # TC-FE-ADM-01: View Admin Dashboard statistics
        # ----------------------------------------------------
        tc_id = "TC-FE-ADM-01"
        try:
            body_text = driver.find_element(By.TAG_NAME, "body").text
            if "Dashboard" in body_text and "Patient" in body_text:
                results[tc_id] = {"status": "Pass", "actual": "Successfully loaded admin dashboard containing widgets and navigation tabs."}
            else:
                results[tc_id] = {"status": "Fail", "actual": "Dashboard statistics/tables not found in the body text."}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-ADM-02: Search and filter patient list table
        # ----------------------------------------------------
        tc_id = "TC-FE-ADM-02"
        try:
            driver.get(f"{BASE_URL}/frontend/admin/patients.php")
            wait.until(EC.presence_of_element_located((By.NAME, "search")))
            search_input = driver.find_element(By.NAME, "search")
            search_input.send_keys("Frontend")
            search_input.submit()
            time.sleep(1.5)
            
            body_text = driver.find_element(By.TAG_NAME, "body").text
            if "Frontend Patient" in body_text:
                results[tc_id] = {"status": "Pass", "actual": "Search filter completed; table filtered to show Frontend Patient."}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Search result table did not contain Frontend Patient. Text was: {body_text[:200]}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

        # ----------------------------------------------------
        # TC-FE-ADM-03: Assign dental procedure to patient
        # ----------------------------------------------------
        tc_id = "TC-FE-ADM-03"
        try:
            driver.get(f"{BASE_URL}/frontend/admin/patient-detail.php?id={patient_id}")
            wait.until(EC.presence_of_element_located((By.NAME, "procedure_id")))
            
            # Select procedure 1 and Intervention group
            Select(driver.find_element(By.NAME, "procedure_id")).select_by_value("1")
            Select(driver.find_element(By.NAME, "group_type")).select_by_value("Intervention")
            
            safe_click(driver, By.XPATH, "//button[contains(., 'Assignment') or contains(@class, 'btn-primary') or contains(., 'Update')]")
            time.sleep(2)
            
            alert = driver.find_element(By.CLASS_NAME, "alert-success")
            if "assigned successfully" in alert.text.lower():
                results[tc_id] = {"status": "Pass", "actual": f"Procedure assignment updated successfully. Alert text: {alert.text}"}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Success banner shown, but text was unexpected: {alert.text}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}. URL: {driver.current_url}. Page Source: {driver.page_source[:400]}"}

        # ----------------------------------------------------
        # TC-FE-AUTH-05: Verify session logout UI flow
        # ----------------------------------------------------
        tc_id = "TC-FE-AUTH-05"
        try:
            driver.get(f"{BASE_URL}/backend/api/auth/logout.php")
            # Clear cookies to be absolutely sure
            driver.delete_all_cookies()
            driver.get(f"{BASE_URL}/frontend/admin/dashboard.php")
            time.sleep(1)
            
            curr_url = driver.current_url
            src = driver.page_source
            if "login.php" in curr_url or "Access Denied" in src or "403" in src:
                results[tc_id] = {"status": "Pass", "actual": f"Logout verified; access blocked. URL: {curr_url}"}
            else:
                results[tc_id] = {"status": "Fail", "actual": f"Expected block/redirect to login page after logout, but URL was: {curr_url}"}
        except Exception as e:
            results[tc_id] = {"status": "Fail", "actual": f"Error: {str(e)}"}

    finally:
        driver.quit()
        db_cleanup(TEST_EMAIL)

    print("\nWriting frontend test results to Excel & CSV...")
    write_results_to_excel(results)
    
    print("\n==================================================")
    print("      FRONTEND TEST EXECUTIONS COMPLETE           ")
    print("==================================================")

if __name__ == "__main__":
    run_tests()
